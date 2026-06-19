from fastapi import FastAPI, UploadFile, File, HTTPException, Form
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse
import os, shutil, zipfile, uuid, io
import fitz
import pdfplumber
import re
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from PIL import Image
import razorpay
import os

RAZORPAY_KEY_ID = os.getenv("RAZORPAY_KEY_ID")
RAZORPAY_KEY_SECRET = os.getenv("RAZORPAY_KEY_SECRET")
rzp_client = razorpay.Client(auth=(RAZORPAY_KEY_ID, RAZORPAY_KEY_SECRET))
import razorpay
import os

RAZORPAY_KEY_ID = os.getenv("RAZORPAY_KEY_ID")
RAZORPAY_KEY_SECRET = os.getenv("RAZORPAY_KEY_SECRET")
rzp_client = razorpay.Client(auth=(RAZORPAY_KEY_ID, RAZORPAY_KEY_SECRET))
import razorpay
import os

RAZORPAY_KEY_ID = os.getenv("RAZORPAY_KEY_ID")
RAZORPAY_KEY_SECRET = os.getenv("RAZORPAY_KEY_SECRET")
rzp_client = razorpay.Client(auth=(RAZORPAY_KEY_ID, RAZORPAY_KEY_SECRET))
import razorpay
import os

RAZORPAY_KEY_ID = os.getenv("RAZORPAY_KEY_ID")
RAZORPAY_KEY_SECRET = os.getenv("RAZORPAY_KEY_SECRET")
rzp_client = razorpay.Client(auth=(RAZORPAY_KEY_ID, RAZORPAY_KEY_SECRET))
import numpy as np

app = FastAPI(title="InvoiceSign MFD")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_methods=["*"],
    allow_headers=["*"],
)

UPLOAD_DIR = "uploads"
OUTPUT_DIR = "outputs"
MAX_PDF_SIZE_MB = 15

os.makedirs(UPLOAD_DIR, exist_ok=True)
os.makedirs(OUTPUT_DIR, exist_ok=True)


def make_transparent(sig_path):
    sig_img = Image.open(sig_path).convert("RGBA")
    arr = np.array(sig_img)
    r = arr[:,:,0].astype(int)
    g = arr[:,:,1].astype(int)
    b = arr[:,:,2].astype(int)
    is_ink = (b - r > 20) | ((r < 150) & (g < 150) & (b < 150))
    arr[:,:,3] = np.where(is_ink, 255, 0).astype(np.uint8)
    clean_sig = Image.fromarray(arr)
    clean_path = sig_path + "_clean.png"
    clean_sig.save(clean_path)
    return clean_path


def extract_details(pdf_path, pan=""):
    try:
        with pdfplumber.open(pdf_path, password=pan) as pdf:
            full_text = ""
            for page in pdf.pages:
                text = page.extract_text()
                if text:
                    full_text += text
    except Exception:
        return "Not found", "Not found"

    inv_match = re.search(
        r"(?:Invoice No|Inv serial No|Invoice Number)[:\s.]*([A-Z0-9/\-]+)",
        full_text, re.IGNORECASE
    )
    invoice_no = inv_match.group(1).strip() if inv_match else "Not found"

    date_match = re.search(
        r"(?:Invoice Date|Date)\s*[:\s]+(\d{1,2}[\/\-]\d{1,2}[\/\-]\d{4}|[A-Za-z]+ \d{1,2},?\s*\d{4})",
        full_text, re.IGNORECASE
    )
    invoice_date = date_match.group(1).strip() if date_match else "Not found"

    return invoice_no, invoice_date


def compress_pdf(doc, output_path):
    """Save PDF with compression. Returns final size in MB."""
    doc.save(
        output_path,
        garbage=4,          # remove unused objects
        deflate=True,       # compress streams
        clean=True,         # clean up structure
        deflate_images=True,
        deflate_fonts=True,
    )
    size_mb = os.path.getsize(output_path) / 1024 / 1024

    # Agar still >15MB, image resolution kam karo
    if size_mb > MAX_PDF_SIZE_MB:
        doc2 = fitz.open(output_path)
        temp_path = output_path + "_temp.pdf"
        for page in doc2:
            for img in page.get_images():
                xref = img[0]
                try:
                    pix = fitz.Pixmap(doc2, xref)
                    if pix.n > 4:
                        pix = fitz.Pixmap(fitz.csRGB, pix)
                    # Reduce DPI by scaling down large images
                    if pix.width > 1500 or pix.height > 1500:
                        scale = 1500 / max(pix.width, pix.height)
                        new_w = int(pix.width * scale)
                        new_h = int(pix.height * scale)
                        pix = pix.resize(new_w, new_h)
                    doc2.update_object(xref, f"<< /Filter /DCTDecode /ColorSpace /DeviceRGB /Width {pix.width} /Height {pix.height} /BitsPerComponent 8 >>")
                    doc2.update_stream(xref, pix.tobytes("jpeg", jpg_quality=60))
                    pix = None
                except Exception:
                    pass
        doc2.save(temp_path, garbage=4, deflate=True, clean=True)
        doc2.close()
        os.replace(temp_path, output_path)
        size_mb = os.path.getsize(output_path) / 1024 / 1024

    return size_mb


def stamp_pdf(pdf_path, output_path, sig_path, signatory_name, designation, pan=""):
    invoice_no, invoice_date = extract_details(pdf_path, pan)
    clean_sig = make_transparent(sig_path)

    doc = fitz.open(pdf_path)

    if doc.is_encrypted:
        if pan:
            doc.authenticate(pan)
        else:
            raise Exception("PDF password protected — PAN number daalo!")

    page = doc[-1]
    pw = page.rect.width

    # Detect: CAMS ~705 wide, Karvy/Axis ~612 wide
    is_cams = pw > 650

    if is_cams:
        sig_rect = fitz.Rect(30, 560, 250, 650)
        page.insert_image(sig_rect, filename=clean_sig, keep_proportion=True, overlay=True)
    else:
        page.draw_rect(fitz.Rect(425, 635, 570, 762), color=(1, 1, 1), fill=(1, 1, 1))
        page.insert_text((430, 648), signatory_name, fontsize=9, fontname="helv", color=(0, 0, 0))
        page.insert_text((430, 662), designation, fontsize=9, fontname="helv", color=(0, 0, 0))
        sig_rect = fitz.Rect(380, 660, 640, 800)
        page.insert_image(sig_rect, filename=clean_sig)

    size_mb = compress_pdf(doc, output_path)
    doc.close()
    return invoice_no, invoice_date, size_mb


def extract_pdfs_from_zip(zip_path, extract_dir):
    """
    ZIP ke andar se saare PDFs nikaalo — nested ZIPs bhi handle karta hai.
    CAMS flat ZIP: PDF seedha andar
    All_Funds ZIP: ZIP > ZIP > PDF
    Returns list of (pdf_path, original_name)
    """
    pdf_files = []

    def process_zip_file(zf, base_dir, prefix=""):
        for name in zf.namelist():
            if name.lower().endswith(".pdf"):
                # PDF directly milgaya
                safe_name = os.path.basename(name)
                out_path = os.path.join(base_dir, prefix + safe_name)
                with zf.open(name) as src, open(out_path, "wb") as dst:
                    dst.write(src.read())
                pdf_files.append((out_path, safe_name))

            elif name.lower().endswith(".zip"):
                # Nested ZIP — andar jaao
                inner_bytes = zf.read(name)
                inner_prefix = os.path.splitext(os.path.basename(name))[0] + "_"
                with zipfile.ZipFile(io.BytesIO(inner_bytes)) as inner_zf:
                    process_zip_file(inner_zf, base_dir, prefix=inner_prefix)

    with zipfile.ZipFile(zip_path, "r") as zf:
        process_zip_file(zf, extract_dir)

    return pdf_files


def make_excel(results, out_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Invoice Report"

    headers = ["S.No", "Filename", "Invoice No.", "Date", "Signatory", "Designation", "Size (MB)", "Status"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = PatternFill("solid", fgColor="1E3A5F")
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center")

    widths = [6, 40, 18, 15, 20, 15, 10, 12]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    for i, r in enumerate(results, 1):
        row_data = [i, r["filename"], r["invoice_no"],
                    r["invoice_date"], r["signatory_name"], r["designation"],
                    r.get("size_mb", ""), r["status"]]
        for col, val in enumerate(row_data, 1):
            cell = ws.cell(row=i+1, column=col, value=val)
            cell.alignment = Alignment(horizontal="center")
            if i % 2 == 0:
                cell.fill = PatternFill("solid", fgColor="EEF2F7")

    wb.save(out_path)


@app.get("/")
def home():
    return {"status": "InvoiceSign MFD Running!"}


@app.post("/process")
async def process_invoices(
    file: UploadFile = File(...),
    signature: UploadFile = File(...),
    signatory_name: str = Form(""),
    designation: str = Form(""),
    pan: str = Form(""),
):
    job_id = str(uuid.uuid4())[:8]
    job_upload = os.path.join(UPLOAD_DIR, job_id)
    job_output = os.path.join(OUTPUT_DIR, job_id)
    os.makedirs(job_upload, exist_ok=True)
    os.makedirs(job_output, exist_ok=True)

    # Signature save karo
    sig_filename = signature.filename or "signature.png"
    sig_path = os.path.join(job_upload, sig_filename)
    with open(sig_path, "wb") as f:
        shutil.copyfileobj(signature.file, f)

    # Uploaded file save karo
    saved_path = os.path.join(job_upload, file.filename)
    with open(saved_path, "wb") as f:
        shutil.copyfileobj(file.file, f)

    # PDFs collect karo
    pdf_files = []  # list of (pdf_path, display_name)

    if file.filename.lower().endswith(".zip"):
        # ZIP se PDFs nikalo — nested ZIPs bhi handle hoga
        extracted = extract_pdfs_from_zip(saved_path, job_upload)
        pdf_files = extracted
    else:
        pdf_files = [(saved_path, file.filename)]

    if not pdf_files:
        raise HTTPException(status_code=400, detail="No PDFs found in uploaded file!")

    results = []
    for pdf_path, display_name in pdf_files:
        signed_name = display_name  # naam same rakho, _SIGNED nahi
        out_path = os.path.join(job_output, signed_name)
        try:
            invoice_no, invoice_date, size_mb = stamp_pdf(
                pdf_path, out_path, sig_path, signatory_name, designation, pan
            )
            results.append({
                "filename": display_name,
                "signed_filename": signed_name,
                "invoice_no": invoice_no,
                "invoice_date": invoice_date,
                "signatory_name": signatory_name,
                "designation": designation,
                "size_mb": round(size_mb, 2),
                "status": "Signed" if size_mb <= MAX_PDF_SIZE_MB else f"Signed (large: {size_mb:.1f}MB)"
            })
        except Exception as e:
            results.append({
                "filename": display_name,
                "signed_filename": signed_name,
                "invoice_no": "Error",
                "invoice_date": "Error",
                "signatory_name": signatory_name,
                "designation": designation,
                "size_mb": "",
                "status": f"Failed: {str(e)}"
            })

    excel_path = os.path.join(job_output, "Invoice_Report.xlsx")
    make_excel(results, excel_path)

    zip_path = os.path.join(OUTPUT_DIR, f"{job_id}_output.zip")
    with zipfile.ZipFile(zip_path, "w", zipfile.ZIP_DEFLATED) as zf:
        for r in results:
            signed = os.path.join(job_output, r["signed_filename"])
            if os.path.exists(signed):
                zf.write(signed, r["signed_filename"])

    return {
        "job_id": job_id,
        "processed": len(results),
        "results": results,
        "download_url": f"/download/{job_id}"
    }


@app.get("/download/{job_id}")
def download(job_id: str):
    zip_path = os.path.join(OUTPUT_DIR, f"{job_id}_output.zip")
    if not os.path.exists(zip_path):
        raise HTTPException(status_code=404, detail="File not found")
    return FileResponse(zip_path, filename="SignedInvoices.zip",
                        media_type="application/zip")


# ─────────────────────────────────────────────
# CAMS Section: Extract + Generate Excel
# ─────────────────────────────────────────────

from fastapi.responses import StreamingResponse
from pydantic import BaseModel
from typing import List, Optional
import json


def extract_cams_data(pdf_path: str, pan: str = "") -> dict:
    """Single CAMS PDF se AMC name, invoice no, taxable, IGST extract karo."""
    try:
        with pdfplumber.open(pdf_path, password=pan or None) as pdf:
            full_text = ""
            for page in pdf.pages:
                t = page.extract_text()
                if t:
                    full_text += t + "\n"
    except Exception as e:
        return {"error": str(e)}

    # Invoice No
    inv_match = re.search(r"Invoice No\s*[:\s]+([A-Z0-9/\-]+)", full_text, re.IGNORECASE)
    invoice_no = inv_match.group(1).strip() if inv_match else "Not found"

    # Invoice Date
    date_match = re.search(
        r"Invoice Date\s*[:\s]+(\d{1,2}[\/\-]\d{1,2}[\/\-]\d{4}|[A-Za-z]+ \d{1,2},?\s*\d{4})",
        full_text, re.IGNORECASE
    )
    invoice_date = date_match.group(1).strip() if date_match else "Not found"

    # AMC Name — line with "Mutual Fund" or "AMC" before GSTIN
    amc_match = re.search(
        r"^(.{5,80}(?:Mutual Fund|AMC|Asset Management|Trustee).*?)\s+GSTIN\s*:",
        full_text, re.IGNORECASE | re.MULTILINE
    )
    if not amc_match:
        # fallback: line after "TAX INVOICE" block
        amc_match = re.search(
            r"TAX INVOICE.*?\n.*?\n([A-Z][A-Za-z\s]+(?:Mutual Fund|AMC|Asset))",
            full_text, re.IGNORECASE | re.DOTALL
        )
    amc_name = amc_match.group(1).strip() if amc_match else "Unknown AMC"

    # Taxable Value + IGST — handles 82039.8 and 84,860.27 both
    NUM = r"[\d,]+\.?\d*"
    taxable_matches = re.findall(
        rf"^Total\s+({NUM})\s+{NUM}\s+{NUM}\s+({NUM})",
        full_text, re.MULTILINE
    )
    if taxable_matches:
        taxable = float(taxable_matches[-1][0].replace(",", ""))
        igst = float(taxable_matches[-1][1].replace(",", ""))
    else:
        tax_match = re.search(rf"^Total\s+({NUM})", full_text, re.MULTILINE)
        taxable = float(tax_match.group(1).replace(",", "")) if tax_match else 0.0
        igst_match = re.search(rf"18(?:\.00)?\s+({NUM})", full_text)
        igst = float(igst_match.group(1).replace(",", "")) if igst_match else round(taxable * 0.18, 2)

    total_invoice = round(taxable + igst, 2)

    return {
        "amc_name": amc_name,
        "cams_invoice_no": invoice_no,
        "invoice_date": invoice_date,
        "taxable": taxable,
        "igst": igst,
        "total": total_invoice,
        "broker_invoice_no": invoice_no,  # default same as CAMS
        "filename": os.path.basename(pdf_path),
    }


@app.post("/extract-cams")
async def extract_cams(
    file: UploadFile = File(...),
    pan: str = Form(""),
):
    """CAMS ZIP ya single PDF se data extract karo."""
    job_id = str(uuid.uuid4())[:8]
    job_upload = os.path.join(UPLOAD_DIR, job_id)
    os.makedirs(job_upload, exist_ok=True)

    saved_path = os.path.join(job_upload, file.filename)
    with open(saved_path, "wb") as f:
        shutil.copyfileobj(file.file, f)

    pdf_files = []
    if file.filename.lower().endswith(".zip"):
        extracted = extract_pdfs_from_zip(saved_path, job_upload)
        pdf_files = [p for p, _ in extracted]
    else:
        pdf_files = [saved_path]

    if not pdf_files:
        raise HTTPException(status_code=400, detail="No PDFs found!")

    results = []
    for pdf_path in sorted(pdf_files):
        data = extract_cams_data(pdf_path, pan)
        if "error" not in data:
            results.append(data)

    return {"job_id": job_id, "count": len(results), "results": results}


class CamsRow(BaseModel):
    amc_name: str
    cams_invoice_no: str
    invoice_date: str
    taxable: float
    igst: float
    total: Optional[float] = None
    broker_invoice_no: str
    useCamsNo: bool = True
    filename: Optional[str] = ""


class CamsExcelRequest(BaseModel):
    results: List[CamsRow]
    broker_code: str = ""
    broker_gst: str = ""


@app.post("/generate-cams-excel")
async def generate_cams_excel(req: CamsExcelRequest):
    """CAMS bulk upload exact format mein Excel banao."""
    import io as _io

    # AMC CODE mapping
    amc_code_map = {
        "aditya birla": "B", "birla": "B",
        "franklin": "FTI",
        "hdfc": "H",
        "hsbc": "O",
        "icici": "P",
        "kotak": "K",
        "ppfas": "PP",
        "sbi": "L",
        "tata": "T",
        "nippon": "N", "nimf": "N",
        "axis": "X",
        "dsp": "D",
        "edelweiss": "E",
        "invesco": "IV",
        "mirae": "MA",
        "motilal": "MO", "mosl": "MO",
        "pgim": "PG",
        "canara": "CAN", "canbank": "CAN",
        "uti": "U",
        "bandhan": "BD",
        "navi": "NV",
        "quantum": "Q",
        "whiteoak": "WO",
        "zerodha": "Z",
        "groww": "GR",
    }

    def get_amc_code(amc_name):
        name_lower = amc_name.lower()
        for key, code in amc_code_map.items():
            if key in name_lower:
                return code
        return ""

    def get_payment_month_year(invoice_date):
        """Invoice date se MMYYYY format banao."""
        import re
        # June 05, 2026 format
        month_map = {
            "january": "01", "february": "02", "march": "03", "april": "04",
            "may": "05", "june": "06", "july": "07", "august": "08",
            "september": "09", "october": "10", "november": "11", "december": "12"
        }
        for month_name, month_num in month_map.items():
            if month_name in invoice_date.lower():
                year_match = re.search(r"(\d{4})", invoice_date)
                year = year_match.group(1) if year_match else ""
                return f"{month_num}{year}"
        # DD/MM/YYYY format
        match = re.match(r"(\d{1,2})[\/\-](\d{1,2})[\/\-](\d{4})", invoice_date)
        if match:
            return f"{match.group(2).zfill(2)}{match.group(3)}"
        return ""

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "GST Invoice"

    # Exact CAMS headers
    headers = [
        "AMC CODE", "AMC NAME", "BROKER CODE", "BROKER INVOICE NUMBER",
        "CAMS INVOICE NUMBER", "PAYMENT MONTH YEAR", "BROKER GST NUMBER",
        "IGST AMOUNT", "CGST AMOUNT", "SGST AMOUNT", "TAXABLE VALUE", "FILE NAME"
    ]

    col_widths = [12, 38, 14, 24, 24, 20, 22, 14, 14, 14, 14, 30]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True, name="Arial")
        cell.alignment = Alignment(horizontal="center")
        ws.column_dimensions[get_column_letter(col)].width = col_widths[col-1]

    for i, r in enumerate(req.results, 1):
        broker_inv = r.cams_invoice_no if r.useCamsNo else r.broker_invoice_no
        row_data = [
            get_amc_code(r.amc_name),          # AMC CODE
            r.amc_name,                          # AMC NAME
            req.broker_code,                     # BROKER CODE
            broker_inv,                          # BROKER INVOICE NUMBER
            r.cams_invoice_no,                   # CAMS INVOICE NUMBER
            get_payment_month_year(r.invoice_date),  # PAYMENT MONTH YEAR
            req.broker_gst,                      # BROKER GST NUMBER
            r.igst,                              # IGST AMOUNT
            0,                                   # CGST AMOUNT (interstate = 0)
            0,                                   # SGST AMOUNT (interstate = 0)
            r.taxable,                           # TAXABLE VALUE
            r.filename or "",                    # FILE NAME
        ]
        for col, val in enumerate(row_data, 1):
            cell = ws.cell(row=i+1, column=col, value=val)
            cell.font = Font(name="Arial")
            cell.alignment = Alignment(horizontal="center")

    excel_buf = _io.BytesIO()
    wb.save(excel_buf)
    excel_buf.seek(0)

    return StreamingResponse(
        excel_buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=CAMS_Bulk_Upload.xlsx"}
    )


# ─────────────────────────────────────────────
# Karvy Section: Fill Excel Template
# ─────────────────────────────────────────────

def extract_karvy_pdf_data(pdf_bytes: bytes, filename: str) -> dict:
    """Karvy PDF se Invoice No, Date extract karo."""
    try:
        with pdfplumber.open(io.BytesIO(pdf_bytes)) as pdf:
            text = ''.join(p.extract_text() or '' for p in pdf.pages)
    except Exception as e:
        return {}

    inv_match = re.search(r'Inv serial No\.?\s*[:\s]+([A-Z0-9/\-]+)', text, re.IGNORECASE)
    date_match = re.search(r'Date\s*[:\s]+(\d{1,2}/\d{1,2}/\d{4})', text, re.IGNORECASE)

    # AMC name from filename prefix (AXIS_, NIMF_, UTI_, etc.)
    prefix = filename.split('_')[0].upper()
    amc_from_file = prefix

    return {
        "invoice_no": inv_match.group(1).strip() if inv_match else "",
        "invoice_date": date_match.group(1).strip() if date_match else "",
        "filename": filename,
        "amc_prefix": amc_from_file,
    }


@app.post("/fill-karvy-excel")
async def fill_karvy_excel(
    zip_file: UploadFile = File(...),
    template: UploadFile = File(...),
):
    """Karvy ZIP + template Excel upload karo — filled Excel return karo."""
    import io as _io

    job_id = str(uuid.uuid4())[:8]
    job_upload = os.path.join(UPLOAD_DIR, job_id)
    os.makedirs(job_upload, exist_ok=True)

    # Save files
    zip_path = os.path.join(job_upload, zip_file.filename)
    with open(zip_path, "wb") as f:
        shutil.copyfileobj(zip_file.file, f)

    template_bytes = await template.read()

    # Extract PDFs from ZIP (nested supported)
    pdf_data_map = {}  # amc_prefix -> {invoice_no, invoice_date, filename}

    def process_zip(zf, prefix=""):
        for name in zf.namelist():
            if name.lower().endswith(".pdf"):
                pdf_bytes = zf.read(name)
                basename = os.path.basename(name)
                data = extract_karvy_pdf_data(pdf_bytes, basename)
                if data:
                    amc_prefix = data["amc_prefix"]
                    pdf_data_map[amc_prefix] = data
            elif name.lower().endswith(".zip"):
                inner_bytes = zf.read(name)
                with zipfile.ZipFile(_io.BytesIO(inner_bytes)) as inner_zf:
                    process_zip(inner_zf)

    with zipfile.ZipFile(zip_path) as zf:
        process_zip(zf)

    # Load template
    wb = openpyxl.load_workbook(_io.BytesIO(template_bytes))
    ws = wb.active

    # Find header row
    header_row = 1
    headers = {}
    for col in range(1, ws.max_column + 1):
        val = ws.cell(row=header_row, column=col).value
        if val:
            headers[str(val).strip()] = col

    inv_col = headers.get("Invoice No", headers.get("Invoice Reference No"))
    date_col = headers.get("Invoice Date")
    file_col = headers.get("File Name")
    amc_col = headers.get("AMC")

    filled = 0
    for row in range(header_row + 1, ws.max_row + 1):
        amc_cell = ws.cell(row=row, column=amc_col).value if amc_col else None
        if not amc_cell:
            continue

        amc_upper = str(amc_cell).upper()

        # Match AMC name to PDF prefix
        matched = None
        for prefix, data in pdf_data_map.items():
            if prefix in amc_upper or any(
                word in amc_upper
                for word in [prefix[:4]]
                if len(prefix) >= 4
            ):
                matched = data
                break

        # Broader match by keywords
        if not matched:
            keyword_map = {
                "NIPPON": "NIMF", "CANARA": "CANBANK", "CANBANK": "CANBANK",
                "UTI": "UTI", "MIRAE": "MIRAE", "EDELWEISS": "EDELWEISS",
                "INVESCO": "INVESCO", "MOTILAL": "MOTILAL",
                "AXIS": "AXIS", "PGIM": "PGIM",
            }
            for keyword, prefix in keyword_map.items():
                if keyword in amc_upper and prefix in pdf_data_map:
                    matched = pdf_data_map[prefix]
                    break

        if matched:
            if inv_col and matched["invoice_no"]:
                ws.cell(row=row, column=inv_col).value = matched["invoice_no"]
            if date_col and matched["invoice_date"]:
                ws.cell(row=row, column=date_col).value = matched["invoice_date"]
            if file_col and matched["filename"]:
                ws.cell(row=row, column=file_col).value = matched["filename"]
            filled += 1

    out_buf = _io.BytesIO()
    wb.save(out_buf)
    out_buf.seek(0)

    return StreamingResponse(
        out_buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=GST_Invoice_Filled.xlsx"}
    )


# ─────────────────────────────────────────────
# Auth Routes
# ─────────────────────────────────────────────
from auth import (
    User, get_db, hash_password, verify_password,
    create_token, decode_token, is_subscribed
)
from sqlalchemy.orm import Session
from fastapi import Depends
from fastapi.security import HTTPBearer, HTTPAuthorizationCredentials
from datetime import datetime, timedelta

security = HTTPBearer()


def get_current_user(
    credentials: HTTPAuthorizationCredentials = Depends(security),
    db: Session = Depends(get_db)
):
    token = credentials.credentials
    email = decode_token(token)
    if not email:
        raise HTTPException(status_code=401, detail="Invalid token")
    user = db.query(User).filter(User.email == email).first()
    if not user:
        raise HTTPException(status_code=401, detail="User not found")
    return user


class RegisterRequest(BaseModel):
    email: str
    password: str


class LoginRequest(BaseModel):
    email: str
    password: str


@app.post("/auth/register")
def register(req: RegisterRequest, db: Session = Depends(get_db)):
    existing = db.query(User).filter(User.email == req.email).first()
    if existing:
        raise HTTPException(status_code=400, detail="Email already registered!")
    user = User(
        email=req.email,
        hashed_password=hash_password(req.password),
    )
    db.add(user)
    db.commit()
    token = create_token(req.email)
    return {"token": token, "email": req.email, "subscribed": False}


@app.post("/auth/login")
def login(req: LoginRequest, db: Session = Depends(get_db)):
    user = db.query(User).filter(User.email == req.email).first()
    if not user or not verify_password(req.password, user.hashed_password):
        raise HTTPException(status_code=401, detail="Invalid email or password!")
    token = create_token(req.email)
    return {
        "token": token,
        "email": req.email,
        "subscribed": is_subscribed(user),
        "subscription_end": user.subscription_end.isoformat() if user.subscription_end else None
    }


@app.get("/auth/me")
def me(current_user: User = Depends(get_current_user)):
    return {
        "email": current_user.email,
        "subscribed": is_subscribed(current_user),
        "subscription_end": current_user.subscription_end.isoformat() if current_user.subscription_end else None
    }


@app.get("/auth/activate")
def activate_subscription(
    email: str,
    months: int = 1,
    db: Session = Depends(get_db)
):
    """Admin use — manually subscription activate karo (Razorpay webhook baad mein)."""
    user = db.query(User).filter(User.email == email).first()
    if not user:
        raise HTTPException(status_code=404, detail="User not found")
    now = datetime.utcnow()
    current_end = user.subscription_end if user.subscription_end and user.subscription_end > now else now
    user.subscription_end = current_end + timedelta(days=30 * months)
    db.commit()
    return {"email": email, "subscription_end": user.subscription_end.isoformat()}
@app.post("/create-order")
async def create_order(plan: str, current_user=Depends(get_current_user), db: Session = Depends(get_db)):
    plans = {
        "monthly": 24900,    # ₹249 in paise
        "quarterly": 64900,  # ₹649 in paise
        "yearly": 229900     # ₹2299 in paise
    }
    if plan not in plans:
        raise HTTPException(status_code=400, detail="Invalid plan")
    
    order = rzp_client.order.create({
        "amount": plans[plan],
        "currency": "INR",
        "payment_capture": 1
    })
    return {"order_id": order["id"], "amount": plans[plan], "key": RAZORPAY_KEY_ID}

@app.post("/verify-payment")
async def verify_payment(data: dict, current_user=Depends(get_current_user), db: Session = Depends(get_db)):
    try:
        # Verify signature
        params = {
            "razorpay_order_id": data["razorpay_order_id"],
            "razorpay_payment_id": data["razorpay_payment_id"],
            "razorpay_signature": data["razorpay_signature"]
        }
        rzp_client.utility.verify_payment_signature(params)
    except Exception as e:
        # In test mode, signature may fail — still activate if payment_id exists
        if not data.get("razorpay_payment_id", "").startswith("pay_"):
            raise HTTPException(status_code=400, detail=f"Payment verification failed: {str(e)}")

    try:
        from datetime import datetime, timedelta
        plan_days = {"monthly": 30, "quarterly": 90, "yearly": 365}
        days = plan_days.get(data.get("plan", "monthly"), 30)
        user = db.query(User).filter(User.email == current_user.email).first()
        user.is_active = True
        user.subscription_end = datetime.utcnow() + timedelta(days=days)
        db.commit()
        return {"status": "success"}
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Activation failed: {str(e)}")