import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime

def generate_report(results, output_path="Invoice_Report.xlsx"):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Invoice Report"

    # Header row
    headers = ["S.No", "Filename", "Invoice No.", "Date", "Signatory", "Designation", "Status"]
    header_fill = PatternFill("solid", fgColor="1E3A5F")
    header_font = Font(bold=True, color="FFFFFF", size=10)

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    # Column widths
    widths = [6, 35, 15, 15, 20, 15, 10]
    for i, width in enumerate(widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width

    # Data rows
    for i, r in enumerate(results, 1):
        row = [
            i,
            r["filename"],
            r["invoice_no"],
            r["invoice_date"],
            "Aniket Chopra",
            "Partner",
            r["status"]
        ]
        for col, val in enumerate(row, 1):
            cell = ws.cell(row=i+1, column=col, value=val)
            cell.alignment = Alignment(horizontal="center")
            if i % 2 == 0:
                cell.fill = PatternFill("solid", fgColor="EEF2F7")

    wb.save(output_path)
    print(f"✅ Report saved: {output_path}")

# Test karo
test_results = [
    {"filename": "Karvy Invoice April'26.pdf", "invoice_no": "28", "invoice_date": "17/04/2026", "status": "Signed"}
]
generate_report(test_results)