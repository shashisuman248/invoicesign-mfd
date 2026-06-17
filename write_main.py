content = '''from fastapi import FastAPI, UploadFile, File, HTTPException, Form
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse
import os, shutil, zipfile, uuid
import fitz
import pdfplumber
import re
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from PIL import Image
import numpy as np

app = FastAPI(title="InvoiceSign MFD")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_methods=["*"],
    allow_headers=["*"],
)

UPLOAD_DIR = "uploads"
OUTPUT_DIR = "outputs"

os.makedirs(UPLOAD_DIR, exist_ok=True)
os.makedirs(OUTPUT_DIR, exist_ok=True)


def make_transparent(sig_path):
    sig_img = Image.open(sig_path).convert("RGBA")
    arr = np.array(sig_img)
    r = arr[:,:,0].astype(int)
    g = arr[:,:,1].astype(int)
    b = arr[:,:,2].astype(int)
    is_ink = (b - r > 20) | ((r < 150) & (g < 150) & (b < 150))
    arr[:,:,3] = np.where(is_ink, 255, 0).astype(np.uint8)
    clean_sig = Image.fromarray(arr)
    clean_path = sig_path + "_clean.png"
    clean_sig.save(clean_path)
    return clean_path


def extract_details(pdf_path):
    with pdfplumber.open(pdf_path) as pdf:
        full_text = ""
        for page in pdf.pages:
            text = page.extract_text()
            if text:
                full_text += text

    inv_match = re.search(
        r"(?:Invoice No|Inv serial No|Invoice Number)[:\\s.]*([A-Z0-9/\\-]+)",
        full_text, re.IGNORECASE
    )
    invoice_no = inv_match.group(1).strip() if inv_match else "Not found"

    date_match = re.search(
        r"(?:Invoice Date|Date)\\s*[:\\s]+(\\d{1,2}[\\/\\-]\\d{1,2}[\\/\\-]\\d{4}|[A-Za-z]+ \\d{1,2},?\\s*\\d{4})",
        full_text, re.IGNORECASE
    )
    invoice_date = date_match.group(1).strip() if date_match else "Not found"

    return invoice_no, invoice_date


def stamp_pdf(pdf_path, output_path, sig_path, signatory_name, designation):
    invoice_no, invoice_date = extract_details(pdf_path)
    clean_sig_path = make_transparent(sig_path)

    doc = fitz.open(pdf_path)
    page = doc[-1]
    pw = page.rect.width
    ph = page.rect.height

    sig_x = pw - 220
    sig_y = ph - 280

    # White box to cover existing printed text
    white_rect = fitz.Rect(sig_x - 60, sig_y - 60, sig_x + 250, sig_y + 160)
    page.draw_rect(white_rect, color=(1, 1, 1), fill=(1, 1, 1))

    # Signatory name
    name_rect = fitz.Rect(sig_x - 50, sig_y - 50, sig_x + 240, sig_y - 35)
    page.insert_textbox(name_rect, signatory_name, fontsize=9,
                        fontname="helv", color=(0, 0, 0), align=0)

    # Designation
    desig_rect = fitz.Rect(sig_x - 50, sig_y - 32, sig_x + 240, sig_y - 17)
    page.insert_textbox(desig_rect, designation, fontsize=8,
                        fontname="helv", color=(0, 0, 0), align=0)

    # Signature image
    sig_rect = fitz.Rect(sig_x, sig_y, sig_x + 300, sig_y + 150)
    page.insert_image(sig_rect, filename=clean_sig_path)

    doc.save(output_path)
    doc.close()
    return invoice_no, invoice_date


def make_excel(results, out_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Invoice Report"

    headers = ["S.No", "Filename", "Invoice No.", "Date", "Signatory", "Designation", "Status"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = PatternFill("solid", fgColor="1E3A5F")
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center")

    widths = [6, 35, 15, 15, 20, 15, 10]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    for i, r in enumerate(results, 1):
        row_data = [i, r["filename"], r["invoice_no"],
                    r["invoice_date"], r["signatory_name"], r["designation"], r["status"]]
        for col, val in enumerate(row_data, 1):
            cell = ws.cell(row=i+1, column=col, value=val)
            cell.alignment = Alignment(horizontal="center")
            if i % 2 == 0:
                cell.fill = PatternFill("solid", fgColor="EEF2F7")

    wb.save(out_path)


@app.get("/")
def home():
    return {"status": "InvoiceSign MFD Running!"}


@app.post("/process")
async def process_invoices(
    file: UploadFile = File(...),
    signature: UploadFile = File(...),
    signatory_name: str = Form("Aniket Chopra"),
    designation: str = Form("Partner"),
):
    job_id = str(uuid.uuid4())[:8]
    job_upload = os.path.join(UPLOAD_DIR, job_id)
    job_output = os.path.join(OUTPUT_DIR, job_id)
    os.makedirs(job_upload, exist_ok=True)
    os.makedirs(job_output, exist_ok=True)

    sig_filename = signature.filename or "signature.png"
    sig_path = os.path.join(job_upload, sig_filename)
    with open(sig_path, "wb") as f:
        shutil.copyfileobj(signature.file, f)

    saved_path = os.path.join(job_upload, file.filename)
    with open(saved_path, "wb") as f:
        shutil.copyfileobj(file.file, f)

    pdf_files = []
    if file.filename.lower().endswith(".zip"):
        with zipfile.ZipFile(saved_path, "r") as zf:
            zf.extractall(job_upload)
        for fname in os.listdir(job_upload):
            if fname.lower().endswith(".pdf"):
                pdf_files.append(os.path.join(job_upload, fname))
    else:
        pdf_files.append(saved_path)

    if not pdf_files:
        raise HTTPException(status_code=400, detail="No PDFs found!")

    results = []
    for pdf_path in pdf_files:
        fname = os.path.basename(pdf_path)
        out_path = os.path.join(job_output, fname.replace(".pdf", "_SIGNED.pdf"))
        try:
            invoice_no, invoice_date = stamp_pdf(pdf_path, out_path, sig_path, signatory_name, designation)
            results.append({
                "filename": fname,
                "invoice_no": invoice_no,
                "invoice_date": invoice_date,
                "signatory_name": signatory_name,
                "designation": designation,
                "status": "Signed"
            })
        except Exception as e:
            results.append({
                "filename": fname,
                "invoice_no": "Error",
                "invoice_date": "Error",
                "signatory_name": signatory_name,
                "designation": designation,
                "status": f"Failed: {str(e)}"
            })

    excel_path = os.path.join(job_output, "Invoice_Report.xlsx")
    make_excel(results, excel_path)

    zip_path = os.path.join(OUTPUT_DIR, f"{job_id}_output.zip")
    with zipfile.ZipFile(zip_path, "w") as zf:
        for r in results:
            signed = os.path.join(job_output, r["filename"].replace(".pdf", "_SIGNED.pdf"))
            if os.path.exists(signed):
                zf.write(signed, r["filename"].replace(".pdf", "_SIGNED.pdf"))
        zf.write(excel_path, "Invoice_Report.xlsx")

    return {
        "job_id": job_id,
        "processed": len(results),
        "results": results,
        "download_url": f"/download/{job_id}"
    }


@app.get("/download/{job_id}")
def download(job_id: str):
    zip_path = os.path.join(OUTPUT_DIR, f"{job_id}_output.zip")
    if not os.path.exists(zip_path):
        raise HTTPException(status_code=404, detail="File not found")
    return FileResponse(zip_path, filename="SignedInvoices.zip",
                        media_type="application/zip")
'''

with open(r'C:\Users\user\OneDrive\Desktop\InvoiceSign\main.py', 'w', encoding='utf-8') as f:
    f.write(content)

print("main.py successfully written!")
